import docx
import openpyxl
from docx.shared import Pt
from docx.oxml.ns import qn
from docx.enum.text import WD_ALIGN_PARAGRAPH
import re

try:
    num = int(input("請輸入審查委員數:"))
    crs = int(input("請輸入模組數:"))
    lea = int(input("請選擇聯盟\n0.健康 1.環境 2.終端(若是健康聯盟請輸入0，以此類推):"))
    res = input("是否已有意見回復?(若是請輸入y，否則輸入n):")
    chk_avd = str(input("是否有利益迴避?(若有請輸入y，否則輸入n):"))

    # 分數及意見總表 (此次寫入目標)
    sc = openpyxl.load_workbook('1-1.xlsx')
    shtsc = sc.worksheets[lea]

    # excel 1-2 from google form審查意見
    gf = openpyxl.load_workbook('1-2.xlsx')
    shtgf = gf.worksheets[0]

    # excel 1 審查委員審查模組表
    if chk_avd=='y':
        avd = openpyxl.load_workbook('1.xlsx')
        shtavd = avd.worksheets[0]

    # 審查委員list，須根據每次審查作調整
    comite = ['吳安宇', '張孟凡', '鄭國興']

    for rev in range(num):
        rev_nam = shtgf.cell(row=rev+2,column=3).value[0:3] # 審查委員名稱
        idx = comite.index(str(rev_nam))+1 # 判斷為委員幾(ex:委員1、委員4)

        for v in range(crs):
            if chk_avd=='y':
                if str(shtavd.cell(row=v+2,column=idx+1).value) == 'x': continue

            # 填分數
            score = shtgf.cell(row=rev+2,column=10+v*7).value
            shtsc.cell(row=v+4, column=5+idx).value = score

            # 審查意見
            point = 1
            opin_fir = shtgf.cell(row=rev+2,column=8+v*7).value #複選題選取部分
            if opin_fir[0]=='無': opin_fir=''
            opin_ls = opin_fir.split(', ')
            opin_fir=''
            for o in opin_ls:
                if o != '' and o[0] != '無':
                    opin_fir = opin_fir + str(point)+'. ' + o + '。\n'
                    point+=1
            # 用換行以及句號 分點呈現
            opin_sec = re.split("[\n|。]", str(shtgf.cell(row=rev+2,column=9+v*7).value)) #打字部分
            opin_final = opin_fir
            for sp in opin_sec:
                if sp != '':
                    opin_final = opin_final + str(point) + '. ' + sp + '。\n'
                    point+=1
            # idx從1開始
            shtsc.cell(row=v+4, column=5+num+1+idx).value = opin_final.strip()

    if res=='y':
        for v in range(crs):
            # 從分數及意見總表取資訊
            mod_num = str(shtsc.cell(row=v+4,column=1).value) # 取模組編號
            mod_nam = shtsc.cell(row=v+4,column=5).value # 取模組名稱
            mod_hos = shtsc.cell(row=v+4,column=4).value # 取模組主持人(教師)
            sch = (shtsc.cell(row=v+4,column=2).value).strip() # 取學校
            dpt = shtsc.cell(row=v+4,column=3).value # 取系所

            doc = docx.Document(str(mod_num)+' '+sch+' '+mod_hos+'_'+mod_nam+'_審查意見回覆.docx')
            tb = doc.tables[0]
            chk=0
            for rev in range(num):
                if chk_avd=='y':
                    if str(shtavd.cell(row=v+2,column=rev+2).value) == 'x':
                        chk+=1
                        continue
                resp = str(tb.rows[rev+1-chk].cells[1].text).strip('委員'+str(rev+1)).strip(':').strip()
                shtsc.cell(row=v+4, column=5+2*num+1+rev+1).value = resp


    sc.save('result.xlsx')
    rs = openpyxl.load_workbook('result.xlsx')
    shtrs = rs.worksheets[lea]

    # 平均分數
    for c in range(crs):
        sum=0.0
        rv_cnt = 0
        for rv in range(num):
            if(shtrs.cell(row=c+4, column=6+rv).value!=None):
                rv_cnt+=1
                sum += float(shtrs.cell(row=c+4, column=6+rv).value)
            else: sum+=0.0
        shtrs.cell(row=c+4, column=6+num).value = sum/rv_cnt
            
    rs.save('result.xlsx')
except Exception as e:
    print(e)
    a=input()
