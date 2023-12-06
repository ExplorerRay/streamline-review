import docx
import openpyxl
from docx.shared import Pt
from docx.oxml.ns import qn
from docx.enum.text import WD_ALIGN_PARAGRAPH


num = int(input("請輸入審查委員數:"))
crs = int(input("請輸入課程數:"))
lea = int(input("請選擇聯盟\n1.健康 2.環境 3.終端(若是健康聯盟請輸入1，以此類推):"))

# 審查委員list，須根據每次審查作調整
comite = ['洪士灝', '馬席彬', '黃世旭', '吳文慶'] 

# excel 1-4 課程大表
full = openpyxl.load_workbook('1-4.xlsx') # source 需修改檔名且放在同資料夾!!
shtfull = full.worksheets[lea-1]

# 目標要填入的excel 1-1 分數及意見總表 output
sc = openpyxl.load_workbook('1-1.xlsx') # source 需修改檔名且放在同資料夾!!
shtsc = sc.worksheets[lea-1]

for v in range(crs):
    # 從課程大表取資訊
    pln_num = str(shtfull.cell(row=v+4,column=2).value)[0:4] # 取計畫編號
    crs_num = shtfull.cell(row=v+4,column=2).value # 取課程編號
    crs_nam = shtfull.cell(row=v+4,column=8).value # 取課程名稱
    key_mod = shtfull.cell(row=v+4,column=15).value # 取重點模組
    crs_hos = shtfull.cell(row=v+4,column=7).value # 取課程主持人(教師)
    sch = shtfull.cell(row=v+4,column=3).value # 取學校
    dpt = shtfull.cell(row=v+4,column=6).value # 取系所

    if shtfull.cell(row=v+4,column=1).value!=None: tbc=0
    # 課程老師已填好回覆的docx檔 input
    doc = docx.Document(str(pln_num)+'案-'+sch+'('+dpt+')_期末審查意見回覆.docx')
    tb = doc.tables[tbc]
    tbc+=1

    for re in range(num):

        # 設定寫入之字型及大小
        doc.styles['Normal'].font.name = "Times New Roman"
        doc.styles['Normal'].font.size = Pt(12)
        doc.styles['Normal']._element.rPr.rFonts.set(qn('w:eastAsia'),'標楷體')

        # 將意見和回復 填入分數意見總表
        opi = str(tb.rows[re+1].cells[0].text).strip('委員'+str(re+1)).strip(':').strip('：').strip()
        resp = str(tb.rows[re+1].cells[1].text).strip('委員'+str(re+1)).strip(':').strip('：').strip()

        # opi_f = ''
        # res_f = ''
        # for o in opi:
        #     if o[0:3]!='委員'+str(re+1):
        #         opi_f = opi_f + o
        # for r in resp:
        #     if r[0:3]!='委員'+str(re+1):
        #         res_f = res_f + r

        shtsc.cell(row=v+4, column=16+re).value = opi
        shtsc.cell(row=v+4, column=20+re).value = resp
        
sc.save('final.xlsx')
