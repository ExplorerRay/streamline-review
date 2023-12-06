import openpyxl

# 分數及意見總表 (此次寫入目標)
sc = openpyxl.load_workbook('1-1.xlsx')
shtsc = sc.worksheets[2] # 2代表終端聯盟

# excel 1-2 from google form審查意見
gf = openpyxl.load_workbook('1-2.xlsx')
shtgf = gf.worksheets[0]

num = int(input("請輸入審查委員數:"))
crs = int(input("請輸入課程數:"))

# 審查委員list，須根據每次審查作調整
comite = ['洪士灝', '馬席彬', '黃世旭', '吳文慶'] 

for re in range(num):
    re_nam = shtgf.cell(row=re+2,column=3).value[0:3] # 審查委員名稱
    idx = comite.index(str(re_nam))+1 # 判斷為委員幾(ex:委員1、委員4)
    for v in range(crs):
        # 填分數
        score = shtgf.cell(row=re+2,column=10+v*7).value
        if score > 4: score = 4
        shtsc.cell(row=v+4, column=10+idx).value = score

        # 審查意見
        opin_fir = shtgf.cell(row=re+2,column=8+v*7).value #複選題選取部分
        if opin_fir[0]=='無': opin_fir=''
        opin_ls = opin_fir.split(', ')
        opin_fir=''
        for o in opin_ls:
            opin_fir = opin_fir + o + '\n'
        opin_sec = str(shtgf.cell(row=re+2,column=9+v*7).value) #打字部分
        shtsc.cell(row=v+4, column=15+idx).value = str(opin_fir + '\n' + opin_sec).strip()

sc.save('result.xlsx')
rs = openpyxl.load_workbook('result.xlsx')
shtrs = rs.worksheets[2] # 2代表終端聯盟

# 平均分數
for c in range(crs):
    sum=0.0
    for rv in range(num):
        # 若沒有所有委員都填好，會出Bug
        if(shtrs.cell(row=c+4, column=11+rv).value!=None):
            sum += float(shtrs.cell(row=c+4, column=11+rv).value)
        else: pass
    shtrs.cell(row=c+4, column=15).value = sum/num
        
rs.save('result.xlsx')

