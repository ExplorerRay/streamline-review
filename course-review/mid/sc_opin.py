import docx
import openpyxl
from docx.shared import Pt
from docx.oxml.ns import qn
from docx.enum.text import WD_ALIGN_PARAGRAPH
import re

num = int(input("請輸入審查委員數:"))
crs = int(input("請輸入課程數:"))
lea = int(input("請選擇聯盟\n1.健康 2.環境 3.終端(若是健康聯盟請輸入1，以此類推):"))
res = input("是否已有意見回復?(若是請輸入y，否則輸入n):")

# 分數及意見總表 (此次寫入目標)
sc = openpyxl.load_workbook('1-1.xlsx')
shtsc = sc.worksheets[lea-1] # 2代表終端聯盟

# excel 1-21 from google form審查意見 上學期
gf1 = openpyxl.load_workbook('1-21.xlsx')
shtgf1 = gf1.worksheets[0]
gf1_cnt=0

# excel 1-22 from google form審查意見 上學期
gf2 = openpyxl.load_workbook('1-22.xlsx')
shtgf2 = gf2.worksheets[0]
gf2_cnt=0


# 審查委員list，須根據每次審查作調整
comite = ['洪士灝', '馬席彬', '黃世旭', '吳文慶'] 

for rev in range(num):
    for v in range(crs):
        smest = str(shtsc.cell(row=v+4,column=10).value)
        # 處理上學期部分
        if smest[-1]=='1':
            re_nam = shtgf1.cell(row=rev+2,column=3).value[0:3] # 審查委員名稱
            idx = comite.index(str(re_nam))+1 # 判斷為委員幾(ex:委員1、委員4)
            score = shtgf1.cell(row=rev+2,column=11+gf1_cnt*8).value
            if score > 4: score = 4
            shtsc.cell(row=v+4, column=10+idx).value = score

            # 審查意見
            point = 1
            opin_fir = shtgf1.cell(row=rev+2,column=9+gf1_cnt*8).value #複選題選取部分
            if opin_fir[0]=='無': opin_fir=''
            opin_ls = opin_fir.split(', ')
            opin_fir=''
            for o in opin_ls:
                if o != '' and o[0] != '無':
                    opin_fir = opin_fir + str(point)+'. ' + o + '。\n'
                    point+=1
            # 用換行以及句號 分點呈現
            opin_sec = re.split("[\n|。]", str(shtgf1.cell(row=rev+2,column=10+gf1_cnt*8).value)) #打字部分
            opin_final = opin_fir
            for sp in opin_sec:
                if sp != '':
                    opin_final = opin_final + str(point) + '. ' + sp + '。\n'
                    point+=1
            shtsc.cell(row=v+4, column=15+idx).value = opin_final.strip()
            gf1_cnt+=1
            
        # 處理下學期部分
        else:
            re_nam = shtgf2.cell(row=rev+2,column=3).value[0:3] # 審查委員名稱
            idx = comite.index(str(re_nam))+1 # 判斷為委員幾(ex:委員1、委員4)
            score = shtgf2.cell(row=rev+2,column=9+gf2_cnt*6).value
            if score > 4: score = 4
            shtsc.cell(row=v+4, column=10+idx).value = score
            # 審查意見
            point = 1
            opin_fir = shtgf2.cell(row=rev+2,column=7+gf2_cnt*6).value #複選題選取部分
            if opin_fir[0]=='無': opin_fir=''
            opin_ls = opin_fir.split(', ')
            opin_fir=''
            for o in opin_ls:
                if o != '' and o[0] != '無':
                    opin_fir = opin_fir + str(point)+'. ' + o + '。\n'
                    point+=1
            # 用換行以及句號 分點呈現
            opin_sec = re.split("[\n|。]", str(shtgf2.cell(row=rev+2,column=8+gf2_cnt*6).value)) #打字部分
            opin_final = opin_fir
            for sp in opin_sec:
                if sp != '':
                    opin_final = opin_final + str(point) + '. ' + sp + '。\n'
                    point+=1
            shtsc.cell(row=v+4, column=15+idx).value = opin_final.strip()
            gf2_cnt+=1

if res=='y':
    for v in range(crs):
        crs_num = shtsc.cell(row=v+4,column=2).value # 取課程編號
        crs_nam = shtsc.cell(row=v+4,column=8).value # 取課程名稱
        crs_hos = shtsc.cell(row=v+4,column=7).value # 取課程主持人(教師)
        sch = shtsc.cell(row=v+4,column=3).value # 取學校
        dpt = shtsc.cell(row=v+4,column=6).value # 取系所
        smest = shtsc.cell(row=v+4,column=10).value # 取開課學期

        doc = docx.Document(str(crs_num)+'課程-'+crs_hos+'老師'+smest+'課程審查意見回覆.docx')
        tb = doc.tables[0]
        for rev in range(num):
            resp = str(tb.rows[rev+1].cells[1].text).strip('委員'+str(rev+1)).strip(':').strip()
            shtsc.cell(row=v+4, column=20+rev).value = resp

sc.save('result.xlsx')
rs = openpyxl.load_workbook('result.xlsx')
shtrs = rs.worksheets[2] # 2代表終端聯盟

# 平均分數
for c in range(crs):
    sum=0.0
    for rv in range(num):
        # 若沒有所有委員都填好，會出Bug
        if(shtrs.cell(row=c+4, column=11+rv).value!=None):
            sum += float(shtrs.cell(row=c+4, column=11+rv).value)
        else: pass
    shtrs.cell(row=c+4, column=15).value = sum/num
        
rs.save('result.xlsx')

