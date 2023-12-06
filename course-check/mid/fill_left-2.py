import openpyxl

crs = int(input("請輸入總課程數:"))
lea = int(input("請選擇聯盟\n1.健康 2.環境 3.終端(若是健康聯盟請輸入1，以此類推):"))

# excel 1-4 課程大表
full = openpyxl.load_workbook('1-4.xlsx') # source 需修改檔名且放在同資料夾!!
shtfull = full.worksheets[lea]

# excel 2-1 目標寫入對象
tg = openpyxl.load_workbook('2-1.xlsx')
shttg = tg.worksheets[0]

skip = 0
for v in range(crs):
    # 從課程大表取資訊
    pln_num = str(shtfull.cell(row=v+5,column=2).value)[0:4] # 取計畫編號
    crs_num = shtfull.cell(row=v+5,column=2).value # 取課程編號
    crs_nam = shtfull.cell(row=v+5,column=15).value # 取課程名稱
    ori_hos = shtfull.cell(row=v+5,column=5).value # 取一開始課程主持人(教師)
    crs_hos = shtfull.cell(row=v+5,column=11).value # 取課程主持人(教師)
    sch = shtfull.cell(row=v+5,column=3).value # 取學校
    ori_dpt = shtfull.cell(row=v+5,column=4).value # 取一開始系所
    dpt = shtfull.cell(row=v+5,column=10).value # 取系所
    smest = shtfull.cell(row=v+5,column=17).value # 取開課學期

    # 若是上學期課程就skip
    if smest[-1]=='1': 
        skip+=1
        continue
    
    # 寫入2-1左側基本資料部分
    shttg.cell(row=5+(v-skip)*3, column=1).value = sch
    shttg.cell(row=5+(v-skip)*3, column=2).value = ori_dpt
    shttg.cell(row=5+(v-skip)*3, column=3).value = ori_hos
    shttg.cell(row=5+(v-skip)*3, column=4).value = dpt
    shttg.cell(row=5+(v-skip)*3, column=5).value = crs_hos
    shttg.cell(row=5+(v-skip)*3, column=6).value = crs_nam
    shttg.cell(row=5+(v-skip)*3, column=7).value = crs_num

tg.save('2-3.xlsx')
